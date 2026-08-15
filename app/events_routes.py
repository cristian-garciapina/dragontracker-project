"""Alliance events (Mobilisation, Seuil de la Guerre) — staff-only routes."""
from io import BytesIO
from datetime import date, time

from typing import Optional
from fastapi import APIRouter, Request, Depends, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .auth import get_db
from .auth import require_staff
from . import queries as q
from .models import Event
from . import bot_client

router = APIRouter(prefix="/staff/events", tags=["events"])
from pathlib import Path
TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"
templates = Jinja2Templates(directory=TEMPLATES_DIR)


@router.get("", response_class=HTMLResponse)
def events_list(
    request: Request,
    season: Optional[int] = None,
    user=Depends(require_staff),
    db: Session = Depends(get_db),
):
    all_seasons = q.list_seasons_for_picker(db)
    season_id_qp = season
    if season_id_qp is not None:
        selected_season = next((s for s in all_seasons if s.id == season_id_qp), None)
        if selected_season is None:
            selected_season = q.get_active_season(db)
    else:
        selected_season = q.get_active_season(db)
    events = q.list_events_for_season(db, selected_season.id) if selected_season else []
    return templates.TemplateResponse("staff/events_list.html", {
        "request": request, "user": user, "season": selected_season,
        "events": events,
        "all_seasons": all_seasons,
        "active_season_id": (q.get_active_season(db).id if q.get_active_season(db) else None),
    })


@router.post("/new")
async def create_event(
    kind: str = Form("ranking"),
    name: str = Form(...),
    date_start: str = Form(""),
    date_end: str = Form(""),
    event_date: str = Form(""),
    event_time: str = Form(""),
    info_url: str = Form(""),
    notes: str = Form(""),
    user=Depends(require_staff),
    db: Session = Depends(get_db),
):
    season = q.get_active_season(db)
    if not season:
        raise HTTPException(400, "No active season")

    kind = (kind or "ranking").strip().lower()
    if kind not in ("ranking", "behemoth", "gate", "spire"):
        raise HTTPException(400, "Invalid kind")

    ds = de = None
    ev_time = None

    if kind == "ranking":
        if not date_start.strip() or not date_end.strip():
            raise HTTPException(400, "Start and end dates required for ranking events")
        try:
            ds = date.fromisoformat(date_start.strip())
            de = date.fromisoformat(date_end.strip())
        except ValueError:
            raise HTTPException(400, "Invalid date format (expected YYYY-MM-DD).")
        if de < ds:
            raise HTTPException(400, "End date must be on or after start date")
    else:
        if not event_date.strip():
            raise HTTPException(400, "Date required")
        if not event_time.strip():
            raise HTTPException(400, "Time required")
        try:
            ds = date.fromisoformat(event_date.strip())
            de = ds
            ev_time = time.fromisoformat(event_time.strip())
        except ValueError:
            raise HTTPException(400, "Invalid date or time format")

    info_url_val = info_url.strip() if kind == "behemoth" and info_url.strip() else None
    notes_val = notes.strip() if kind != "ranking" and notes.strip() else None

    ev = Event(
        name=name.strip(),
        kind=kind,
        date_start=ds,
        date_end=de,
        event_time=ev_time,
        info_url=info_url_val,
        notes=notes_val,
        season_id=season.id,
        created_by=getattr(user, "username", None),
    )
    db.add(ev)
    db.commit()

    if kind == "ranking":
        q.prefill_event_eligibility(db, ev.id, season.id)

    await bot_client.notify_new_event({
        "event_id": ev.id,
        "kind": ev.kind,
        "name": ev.name,
        "date_start": ev.date_start.isoformat(),
        "date_end": ev.date_end.isoformat(),
        "event_time": ev.event_time.strftime("%H:%M") if ev.event_time else None,
        "info_url": ev.info_url,
        "notes": ev.notes,
        "season_id": ev.season_id,
        "created_by": ev.created_by,
    })
    return RedirectResponse(f"/staff/events/{ev.id}", status_code=303)


@router.get("/{event_id}", response_class=HTMLResponse)
def event_detail(event_id: int, request: Request,
                 user=Depends(require_staff), db: Session = Depends(get_db)):
    ev = q.get_event(db, event_id)
    if not ev:
        raise HTTPException(404)
    participations = q.get_event_participations_ranked(db, event_id)
    zero_count = sum(1 for p in participations if p["points"] == 0)
    eligibility_rows = q.list_event_eligibility(db, event_id) if ev.kind == "ranking" else []
    eligible_total = sum(1 for r in eligibility_rows if r["is_eligible"])
    rsvps = q.get_event_rsvps(db, event_id)
    return templates.TemplateResponse("staff/event_detail.html", {
        "request": request, "user": user, "event": ev,
        "participations": participations,
        "eligibility_rows": eligibility_rows,
        "eligible_total": eligible_total,
        "zero_count": zero_count,
        "rsvps": rsvps,
    })


@router.get("/{event_id}/json")
def event_detail_json(event_id: int,
                      user=Depends(require_staff), db: Session = Depends(get_db)):
    ev = q.get_event(db, event_id)
    if not ev:
        raise HTTPException(404)
    participations = q.get_event_participations_ranked(db, event_id)
    zero_count = sum(1 for p in participations if p["points"] == 0)
    eligibility_rows = q.list_event_eligibility(db, event_id) if ev.kind == "ranking" else []
    eligible_total = sum(1 for r in eligibility_rows if r["is_eligible"])
    rsvps = q.get_event_rsvps(db, event_id)
    def _iso(dt):
        return dt.isoformat() if dt else None
    return {
        "coming": [{"character_id": r["character_id"], "name": r["name"],
                    "responded_at": _iso(r["responded_at"])} for r in rsvps["yes"]],
        "not_coming": [{"character_id": r["character_id"], "name": r["name"],
                        "responded_at": _iso(r["responded_at"])} for r in rsvps["no"]],
        "participations": [{"rank": p["rank"], "character_id": p["character_id"],
                            "name": p["name"], "points": p["points"],
                            "notes": p.get("notes")} for p in participations],
        "counts": {
            "coming": len(rsvps["yes"]),
            "not_coming": len(rsvps["no"]),
            "participants": len(participations),
            "zero": zero_count,
            "eligible": eligible_total,
        },
    }


@router.post("/{event_id}/import")
async def import_excel(
    event_id: int,
    file: UploadFile = File(...),
    user=Depends(require_staff),
    db: Session = Depends(get_db),
):
    ev = q.get_event(db, event_id)
    if not ev:
        raise HTTPException(404)
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(400, "Invalid xlsx")
    ws = wb.active
    rows = []
    parse_errors = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            cid = int(row[0])
            pts = int(row[1]) if len(row) > 1 and row[1] is not None else 0
        except (ValueError, TypeError):
            parse_errors += 1
            continue
        note = None
        if len(row) > 2 and row[2] is not None:
            s = str(row[2]).strip()
            note = s or None
        rows.append((cid, pts, note))
    inserted, updated, skipped = q.upsert_event_participations(db, event_id, rows)
    msg = f"Imported {inserted} new · updated {updated} · skipped {skipped} unknown"
    if parse_errors:
        msg += f" · {parse_errors} parse error(s)"
    return RedirectResponse(f"/staff/events/{event_id}?flash={msg}", status_code=303)


@router.post("/{event_id}/manual")
async def save_manual(
    event_id: int, request: Request,
    user=Depends(require_staff), db: Session = Depends(get_db),
):
    if not q.get_event(db, event_id):
        raise HTTPException(404)
    form = await request.form()
    rows = []
    for key, value in form.multi_items():
        if not key.startswith("pts_"):
            continue
        try:
            cid = int(key[4:])
            pts = int(value) if value else 0
        except ValueError:
            continue
        rows.append((cid, pts, None))
    inserted, updated, _ = q.upsert_event_participations(db, event_id, rows)
    return RedirectResponse(
        f"/staff/events/{event_id}?flash=Saved {inserted + updated} row(s)",
        status_code=303,
    )


@router.post("/{event_id}/delete")
async def delete_event_route(event_id: int,
                             user=Depends(require_staff), db: Session = Depends(get_db)):
    ev = q.get_event(db, event_id)
    if ev is None:
        return RedirectResponse("/staff/events", status_code=303)
    # Capture Discord message coords before deletion
    msg_id = ev.discord_message_id
    ch_id = ev.discord_channel_id
    q.delete_event(db, event_id)
    if msg_id and ch_id:
        await bot_client.notify_event_deleted({
            "message_id": msg_id,
            "channel_id": ch_id,
        })
    return RedirectResponse("/staff/events", status_code=303)


@router.post("/{event_id}/eligibility")
async def update_eligibility(
    event_id: int, request: Request,
    user=Depends(require_staff), db: Session = Depends(get_db),
):
    if not q.get_event(db, event_id):
        raise HTTPException(404)
    form = await request.form()
    eligible_ids: set[int] = set()
    for key in form.keys():
        if not key.startswith("elig_"):
            continue
        try:
            eligible_ids.add(int(key[5:]))
        except ValueError:
            continue
    on, off = q.set_event_eligibility(db, event_id, eligible_ids)
    msg = f"Marked {on} eligible, {off} not eligible"
    return RedirectResponse(f"/staff/events/{event_id}?flash={msg}", status_code=303)



@router.post("/{event_id}/add-participant")
async def add_manual_participant(
    event_id: int,
    character_id: int = Form(...),
    in_game_name: str = Form(...),
    points: int = Form(0),
    notes: str = Form(""),
    user=Depends(require_staff),
    db: Session = Depends(get_db),
):
    if not q.get_event(db, event_id):
        raise HTTPException(404)
    if character_id <= 0:
        raise HTTPException(400, "Invalid character_id")
    name = in_game_name.strip()
    if not name:
        raise HTTPException(400, "In-game name required")
    ghost_created = q.ensure_member_exists(db, character_id, name)
    action = q.add_manual_participation(
        db, event_id, character_id, max(0, points), notes.strip() or None
    )
    if ghost_created:
        msg = f"Ghost member #{character_id} created and added ({points} pts)"
    elif action == "created":
        msg = f"Added participation for #{character_id} ({points} pts)"
    else:
        msg = f"Updated participation for #{character_id} ({points} pts)"
    return RedirectResponse(f"/staff/events/{event_id}?flash={msg}", status_code=303)
